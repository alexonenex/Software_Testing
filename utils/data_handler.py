"""
数据处理工具模块
支持 YAML / Excel 数据驱动
"""
import os
import yaml
from utils.logger import log


def load_yaml(file_path: str) -> dict | list:
    """
    加载 YAML 文件
    :param file_path: YAML 文件路径
    :return: 解析后的数据
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"数据文件不存在: {file_path}")

    with open(file_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    log.debug(f"已加载 YAML 数据: {file_path}")
    return data


def load_excel(file_path: str, sheet_name: str = None) -> list[dict]:
    """
    加载 Excel 文件
    :param file_path: Excel 文件路径
    :param sheet_name: 工作表名称，默认取第一个
    :return: 字典列表，每行一个字典
    """
    from openpyxl import load_workbook

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"数据文件不存在: {file_path}")

    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 首行为表头
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)

    log.debug(f"已加载 Excel 数据: {file_path} (行数: {len(data)})")
    return data
